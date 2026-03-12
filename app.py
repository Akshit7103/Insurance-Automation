import os
import csv
import uuid
import json
import time
import shutil
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter

app = FastAPI(title="Insurance Automation - Protiviti")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_DIR = os.path.join(BASE_DIR, "temp_files")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
os.makedirs(TEMP_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════
# Column Mappings
# ══════════════════════════════════════════════════════════════

# --- GLA ---
GLA_INPUT_COLS = {
    "Status":       "B",
    "CNTTYPE":      "Y",
    "PREM_TERM":    "AC",
    "PREMCESDTE":   "AH",
    "ORIGINAL_SA":  "AS",
    "NEXT_PAYDATE": "AW",
}
GLA_OUTPUT_COL = "BA"

# --- FRA ---
FRA_INPUT_COLS = {
    "Status":       "B",
    "CNTTYPE":      "Y",
    "RISK_TERM":    "AB",
    "PREM_TERM":    "AC",
    "FREQUENCY":    "AD",
    "PREMCESDTE":   "AH",
    "RCDDATE":      "AE",
    "PAIDTODATE":   "AF",
    "ORIGINAL_SA":  "AS",
    "NEXT_PAYDATE": "AW",
    "BASE_PREMIUM": "AZ",
}
FRA_OUTPUT_COL = "BF"

# Frequency → annualization divisor
FREQ_DIVISOR = {1: 1.0, 2: 0.51, 4: 0.26, 12: 0.09083}

# ══════════════════════════════════════════════════════════════
# Excel Styling
# ══════════════════════════════════════════════════════════════

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

OUTPUT_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="000000")
INPUT_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
VALUE_FONT = Font(name="Calibri", size=11, color="1F4E79")
VALUE_FONT_HIGHLIGHT = Font(name="Calibri", bold=True, size=11, color="006100")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)
OUTPUT_HEADER_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
    top=Side(style="medium", color="C8A53A"), bottom=Side(style="medium", color="C8A53A"),
)
DATA_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"),
)
INPUT_HEADER_BORDER = Border(
    left=Side(style="thin", color="8EA9C8"), right=Side(style="thin", color="8EA9C8"),
    top=Side(style="medium", color="1F4E79"), bottom=Side(style="medium", color="1F4E79"),
)


# ══════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════

def csv_to_xlsx(csv_path: str, xlsx_path: str):
    wb = Workbook()
    ws = wb.active
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(xlsx_path)
    wb.close()


def cleanup_old_files(max_age_seconds: int = 3600):
    now = time.time()
    for fname in os.listdir(TEMP_DIR):
        fpath = os.path.join(TEMP_DIR, fname)
        if os.path.isfile(fpath) and (now - os.path.getmtime(fpath)) > max_age_seconds:
            try:
                os.remove(fpath)
            except OSError:
                pass


def sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data)}\n\n"


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def years_between(start, end):
    """Return complete years between two dates (like Excel DATEDIF 'Y')."""
    if start is None or end is None:
        return 0
    years = end.year - start.year
    if (end.month, end.day) < (start.month, start.day):
        years -= 1
    return max(years, 0)


def fmt_date(d):
    if isinstance(d, datetime):
        return d.strftime("%d-%b-%Y")
    return str(d) if d else "-"


# ══════════════════════════════════════════════════════════════
# Formatting
# ══════════════════════════════════════════════════════════════

def format_entire_sheet(ws, header_row, data_start, max_row, output_col):
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        if c == output_col:
            continue
        if cell.value is not None:
            cell.font = INPUT_HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = INPUT_HEADER_BORDER
    for row_num in range(data_start, max_row + 1):
        for c in range(1, max_col + 1):
            if c == output_col:
                continue
            cell = ws.cell(row=row_num, column=c)
            cell.border = DATA_BORDER
            cell.alignment = Alignment(vertical="center")
    for c in range(1, max_col + 1):
        if c == output_col:
            continue
        cl = get_column_letter(c)
        mx = 0
        for r in range(header_row, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                mx = max(mx, len(str(v)))
        ws.column_dimensions[cl].width = min(max(mx + 3, 8), 30)


def format_output_column(ws, header_row, data_start, max_row, out_col, header_text):
    hc = ws.cell(row=header_row, column=out_col)
    hc.value = header_text
    hc.fill = YELLOW_FILL
    hc.font = OUTPUT_HEADER_FONT
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hc.border = OUTPUT_HEADER_BORDER
    ws.column_dimensions[get_column_letter(out_col)].width = 24
    for r in range(data_start, max_row + 1):
        cell = ws.cell(row=r, column=out_col)
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = THIN_BORDER
        if cell.value is not None and cell.value != 0:
            cell.fill = LIGHT_GREEN_FILL
            cell.font = VALUE_FONT_HIGHLIGHT
        else:
            cell.fill = LIGHT_YELLOW_FILL
            cell.font = VALUE_FONT


# ══════════════════════════════════════════════════════════════
# GLA Processing
# ══════════════════════════════════════════════════════════════

def process_gla_stream(input_path, output_path, header_row, file_id, original_filename):
    try:
        col = {n: column_index_from_string(l) for n, l in GLA_INPUT_COLS.items()}
        out_col = column_index_from_string(GLA_OUTPUT_COL)

        yield sse("stage", {"stage": "reading"})
        wb = load_workbook(input_path)
        ws = wb.active
        total_rows = max(0, ws.max_row - header_row)

        yield sse("stage", {"stage": "processing", "total": total_rows})
        data_start = header_row + 1
        processed = skipped = 0
        total_val = 0.0
        preview = []
        batch = max(1, total_rows // 200)

        for rn in range(data_start, ws.max_row + 1):
            status      = ws.cell(row=rn, column=col["Status"]).value
            cnttype     = ws.cell(row=rn, column=col["CNTTYPE"]).value
            prem_term   = ws.cell(row=rn, column=col["PREM_TERM"]).value
            premcesdte  = ws.cell(row=rn, column=col["PREMCESDTE"]).value
            original_sa = ws.cell(row=rn, column=col["ORIGINAL_SA"]).value
            next_paydate = ws.cell(row=rn, column=col["NEXT_PAYDATE"]).value

            if status is None and cnttype is None and original_sa is None:
                skipped += 1
            else:
                gla = 0.0
                s = str(status).strip().upper() if status else ""
                c = str(cnttype).strip().upper() if cnttype else ""
                if s == "IF" and c in ("MSB", "SMB"):
                    try:
                        sa = float(original_sa) if original_sa is not None else 0
                        pt = float(prem_term) if prem_term is not None else 0
                        gla = round(0.01 * sa * pt, 2)
                    except (ValueError, TypeError):
                        gla = 0.0
                ws.cell(row=rn, column=out_col).value = gla
                processed += 1
                total_val += gla
                if len(preview) < 25:
                    preview.append({
                        "row": rn, "status": str(status) if status else "-",
                        "cnttype": str(cnttype) if cnttype else "-",
                        "prem_term": prem_term if prem_term is not None else "-",
                        "premcesdte": fmt_date(premcesdte),
                        "original_sa": original_sa if original_sa is not None else "-",
                        "next_paydate": fmt_date(next_paydate),
                        "result": gla,
                    })

            cur = processed + skipped
            if cur % batch == 0 or rn == ws.max_row:
                yield sse("progress", {
                    "current": cur, "total": total_rows,
                    "processed": processed, "skipped": skipped,
                    "result_total": round(total_val, 2),
                    "percent": round(cur / total_rows * 100, 1) if total_rows > 0 else 100,
                })

        format_entire_sheet(ws, header_row, data_start, ws.max_row, out_col)
        format_output_column(ws, header_row, data_start, ws.max_row, out_col, "Protiviti Output GLA")

        yield sse("stage", {"stage": "saving"})
        wb.save(output_path)
        wb.close()
        try:
            if os.path.exists(input_path): os.remove(input_path)
        except OSError:
            pass

        out_name = original_filename.rsplit(".", 1)[0] + "_GLA_output.xlsx" if original_filename else "output.xlsx"
        yield sse("complete", {
            "file_id": file_id, "output_filename": out_name,
            "total_rows": total_rows, "processed": processed, "skipped": skipped,
            "result_total": round(total_val, 2), "preview": preview,
        })
    except Exception as e:
        for p in [input_path, output_path]:
            if os.path.exists(p):
                try: os.remove(p)
                except OSError: pass
        yield sse("error", {"message": str(e)})


# ══════════════════════════════════════════════════════════════
# FRA Processing
# ══════════════════════════════════════════════════════════════

def process_fra_stream(input_path, output_path, header_row, file_id, original_filename):
    try:
        col = {n: column_index_from_string(l) for n, l in FRA_INPUT_COLS.items()}
        out_col = column_index_from_string(FRA_OUTPUT_COL)

        yield sse("stage", {"stage": "reading"})
        wb = load_workbook(input_path)
        ws = wb.active
        total_rows = max(0, ws.max_row - header_row)

        yield sse("stage", {"stage": "processing", "total": total_rows})
        data_start = header_row + 1
        processed = skipped = 0
        total_val = 0.0
        preview = []
        batch = max(1, total_rows // 200)

        for rn in range(data_start, ws.max_row + 1):
            status       = ws.cell(row=rn, column=col["Status"]).value
            cnttype      = ws.cell(row=rn, column=col["CNTTYPE"]).value
            risk_term    = ws.cell(row=rn, column=col["RISK_TERM"]).value
            prem_term    = ws.cell(row=rn, column=col["PREM_TERM"]).value
            frequency    = ws.cell(row=rn, column=col["FREQUENCY"]).value
            premcesdte   = to_date(ws.cell(row=rn, column=col["PREMCESDTE"]).value)
            rcddate      = to_date(ws.cell(row=rn, column=col["RCDDATE"]).value)
            paidtodate   = to_date(ws.cell(row=rn, column=col["PAIDTODATE"]).value)
            original_sa  = ws.cell(row=rn, column=col["ORIGINAL_SA"]).value
            next_paydate = to_date(ws.cell(row=rn, column=col["NEXT_PAYDATE"]).value)
            base_premium = ws.cell(row=rn, column=col["BASE_PREMIUM"]).value

            if status is None and cnttype is None and original_sa is None:
                skipped += 1
            else:
                fra = 0.0
                s = str(status).strip().upper() if status else ""
                c = str(cnttype).strip().upper() if cnttype else ""

                try:
                    rt = int(risk_term) if risk_term is not None else 0
                    freq = int(frequency) if frequency is not None else 1
                    bp = float(base_premium) if base_premium is not None else 0

                    # Step 1: Annualized Premium
                    divisor = FREQ_DIVISOR.get(freq, 1.0)
                    ap = bp / divisor if divisor != 0 else 0

                    # Step 2: Policy Year Check
                    policy_year_check = rt - years_between(rcddate, next_paydate)

                    # Step 3: Years premium paid
                    years_paid = years_between(rcddate, paidtodate)

                    # Step 4: Paid up factor
                    puf = 0.0
                    if premcesdte and rcddate and (premcesdte - rcddate).days != 0 and paidtodate:
                        puf = round((paidtodate - rcddate).days / (premcesdte - rcddate).days, 2)

                    # Step 5: FRA
                    if c == "FSP" and policy_year_check == 1:
                        if s in ("DH", "SU", "CF"):
                            fra = 0.0
                        elif s == "IF":
                            fra = ap * 0.08 + ap * 0.09 + ap * 0.10 * (rt - 2)
                        else:
                            fra = (ap * 0.08 + ap * 0.09
                                   + ap * 0.10 * (years_paid - 2)
                                   + ap * puf * (rt - years_paid) * 0.10)
                    fra = round(fra, 2)
                except (ValueError, TypeError):
                    fra = 0.0

                ws.cell(row=rn, column=out_col).value = fra
                processed += 1
                total_val += fra
                if len(preview) < 25:
                    preview.append({
                        "row": rn, "status": str(status) if status else "-",
                        "cnttype": str(cnttype) if cnttype else "-",
                        "risk_term": risk_term if risk_term is not None else "-",
                        "frequency": frequency if frequency is not None else "-",
                        "base_premium": base_premium if base_premium is not None else "-",
                        "rcddate": fmt_date(rcddate),
                        "paidtodate": fmt_date(paidtodate),
                        "result": fra,
                    })

            cur = processed + skipped
            if cur % batch == 0 or rn == ws.max_row:
                yield sse("progress", {
                    "current": cur, "total": total_rows,
                    "processed": processed, "skipped": skipped,
                    "result_total": round(total_val, 2),
                    "percent": round(cur / total_rows * 100, 1) if total_rows > 0 else 100,
                })

        format_entire_sheet(ws, header_row, data_start, ws.max_row, out_col)
        format_output_column(ws, header_row, data_start, ws.max_row, out_col, "Protiviti Output FRA")

        yield sse("stage", {"stage": "saving"})
        wb.save(output_path)
        wb.close()
        try:
            if os.path.exists(input_path): os.remove(input_path)
        except OSError:
            pass

        out_name = original_filename.rsplit(".", 1)[0] + "_FRA_output.xlsx" if original_filename else "output.xlsx"
        yield sse("complete", {
            "file_id": file_id, "output_filename": out_name,
            "total_rows": total_rows, "processed": processed, "skipped": skipped,
            "result_total": round(total_val, 2), "preview": preview,
        })
    except Exception as e:
        for p in [input_path, output_path]:
            if os.path.exists(p):
                try: os.remove(p)
                except OSError: pass
        yield sse("error", {"message": str(e)})


# ══════════════════════════════════════════════════════════════
# Routes
# ══════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
async def serve_index():
    with open(os.path.join(TEMPLATES_DIR, "index.html"), "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())


@app.post("/api/process")
async def api_process(
    file: UploadFile = File(...),
    header_row: int = Form(default=2),
    calc_type: str = Form(default="gla"),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Please upload an Excel (.xlsx) or CSV (.csv) file")
    if calc_type not in ("gla", "fra"):
        raise HTTPException(status_code=400, detail="Invalid calculation type")

    cleanup_old_files()
    file_id = str(uuid.uuid4())
    is_csv = file.filename.lower().endswith(".csv")
    raw_ext = ".csv" if is_csv else ".xlsx"
    raw_path = os.path.join(TEMP_DIR, f"{file_id}_raw{raw_ext}")
    input_path = os.path.join(TEMP_DIR, f"{file_id}_in.xlsx")
    output_path = os.path.join(TEMP_DIR, f"{file_id}_out.xlsx")

    try:
        with open(raw_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")

    if is_csv:
        try:
            csv_to_xlsx(raw_path, input_path)
            os.remove(raw_path)
        except Exception as e:
            for p in [raw_path, input_path]:
                if os.path.exists(p):
                    os.remove(p)
            raise HTTPException(status_code=400, detail=f"Failed to read CSV: {str(e)}")
    else:
        os.rename(raw_path, input_path)

    stream_fn = process_fra_stream if calc_type == "fra" else process_gla_stream
    return StreamingResponse(
        stream_fn(input_path, output_path, header_row, file_id, file.filename),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )


@app.get("/api/download/{file_id}")
async def api_download(file_id: str, filename: str = "output.xlsx"):
    output_path = os.path.join(TEMP_DIR, f"{file_id}_out.xlsx")
    if not os.path.exists(output_path):
        raise HTTPException(status_code=404, detail="File not found or has expired")
    return FileResponse(output_path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
